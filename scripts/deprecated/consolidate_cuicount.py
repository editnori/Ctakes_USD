#!/usr/bin/env python3
# DEPRECATED: Prefer Java workbook builder (tools/reporting/ExcelXmlReport.java) and CuiCounts sheet.
from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple


def find_cuicount_files(root: Path) -> List[Path]:
    return [p for p in root.rglob("*.cuicount.bsv") if p.is_file()]


@dataclass
class CuiRow:
    doc: str
    cui: str
    negated: Optional[bool]
    count: int
    pipeline: str


def _parse_bool(val: str) -> Optional[bool]:
    if val is None:
        return None
    t = val.strip().lower()
    if t in {"1", "true", "t", "yes", "y", "neg", "negated"}:
        return True
    if t in {"0", "false", "f", "no", "n", "pos", "positive"}:
        return False
    try:
        n = int(t)
        if n < 0:
            return True
        if n > 0:
            return False
    except Exception:
        pass
    return None


def _looks_like_cui(token: str) -> bool:
    return bool(re.fullmatch(r"C\d{2,8}", token.strip(), flags=re.IGNORECASE))


def _infer_columns(header_cells: List[str]) -> Dict[str, int]:
    idx: Dict[str, int] = {}
    for i, h in enumerate(header_cells):
        k = h.strip().lower()
        if k == "cui" and "cui" not in idx:
            idx["cui"] = i
        elif k in {"negated", "negation", "polarity"} and "negated" not in idx:
            idx["negated"] = i
        elif k in {"count", "freq", "frequency"} and "count" not in idx:
            idx["count"] = i
        elif k in {"doc", "doc_id", "document", "document_id", "file", "filename"} and "doc" not in idx:
            idx["doc"] = i
    return idx


def _detect_pipeline_for_file(fp: Path, root: Path, fallback: Optional[str]) -> str:
    if fallback:
        return fallback
    parts = [p.name for p in fp.resolve().parents]
    parts.append(fp.name)
    for name in parts:
        m = re.match(r"^S_([A-Za-z0-9_]+)", name)
        if m:
            return m.group(1)
    for name in parts:
        for token in (
            "TsSectionedTemporalCoref",
            "WSD_Compare",
            "core",
            "temp_coref",
        ):
            if token in name:
                return token
    try:
        rel = fp.resolve().relative_to(root.resolve())
        if rel.parts:
            return rel.parts[0]
    except Exception:
        pass
    return "default"


def read_cuicount_file(path: Path, root: Path, explicit_pipeline: Optional[str]) -> Iterable[CuiRow]:
    doc_id = path.stem.replace(".cuicount", "")
    pipeline = _detect_pipeline_for_file(path, root, explicit_pipeline)
    with path.open("r", encoding="utf-8", errors="ignore") as f:
        first = f.readline()
        if not first:
            return []
        sep = "|"
        first_cells = [c.strip() for c in first.rstrip("\n\r").split(sep)]
        has_header = any(re.search(r"[A-Za-z]", c) for c in first_cells)
        col_map: Dict[str, int] = {}
        data_lines: List[List[str]] = []
        if has_header and ("cui" in {c.lower() for c in first_cells} or "count" in {c.lower() for c in first_cells}):
            col_map = _infer_columns(first_cells)
        else:
            data_lines.append(first_cells)
        for line in f:
            if not line.strip():
                continue
            data_lines.append([c.strip() for c in line.rstrip("\n\r").split(sep)])

    if not col_map:
        sample = data_lines[0]
        cui_idx = next((i for i, v in enumerate(sample) if _looks_like_cui(v)), 0)
        cand_counts = [i for i, v in enumerate(sample) if v.isdigit()]
        count_idx = cand_counts[-1] if cand_counts else None
        neg_idx = None
        for i, v in enumerate(sample):
            if _parse_bool(v) is not None:
                neg_idx = i
                break
        col_map = {"cui": cui_idx}
        if count_idx is not None:
            col_map["count"] = count_idx
        if neg_idx is not None:
            col_map["negated"] = neg_idx

    results: List[CuiRow] = []
    for cells in data_lines:
        try:
            cui = cells[col_map.get("cui", 0)].strip()
        except Exception:
            continue
        negated: Optional[bool] = None
        if "negated" in col_map and col_map["negated"] < len(cells):
            negated = _parse_bool(cells[col_map["negated"]])
        count = 1
        if "count" in col_map and col_map["count"] < len(cells):
            try:
                count = int(cells[col_map["count"]])
            except Exception:
                pass
        doc = (
            cells[col_map["doc"]].strip()
            if "doc" in col_map and col_map["doc"] < len(cells)
            else doc_id
        )
        results.append(CuiRow(doc=doc, cui=cui, negated=negated, count=count, pipeline=pipeline))
    return results


def aggregate(rows: Iterable[CuiRow]) -> List[Tuple[str, Optional[bool], int]]:
    agg: Dict[Tuple[str, Optional[bool]], int] = defaultdict(int)
    for r in rows:
        agg[(r.cui, r.negated)] += r.count
    items = [(cui, neg, cnt) for (cui, neg), cnt in agg.items()]
    items.sort(key=lambda x: (-x[2], x[0], (1 if x[1] else 0) if x[1] is not None else 2))
    return items


def aggregate_by_pipeline(rows: Iterable[CuiRow]) -> Dict[str, List[Tuple[str, Optional[bool], int]]]:
    buckets: Dict[str, List[CuiRow]] = defaultdict(list)
    for r in rows:
        buckets[r.pipeline].append(r)
    return {p: aggregate(rs) for p, rs in buckets.items()}


def write_csv(path: Path, headers: List[str], rows: List[Tuple]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in rows:
            w.writerow(list(r))


def _parse_color_hex(s: Optional[str]) -> Optional[str]:
    if not s:
        return None
    ss = s.strip().lstrip("#").upper()
    if re.fullmatch(r"[0-9A-F]{6}", ss):
        return ss
    return None


def write_xlsx(
    xlsx_path: Path,
    agg_rows: List[Tuple[str, Optional[bool], int]],
    per_doc_rows: Optional[List[Tuple[str, str, Optional[bool], int]]] = None,
    pipeline_name: Optional[str] = None,
    header_color: Optional[str] = None,
    per_pipeline_aggs: Optional[Dict[str, List[Tuple[str, Optional[bool], int]]]] = None,
    per_pipeline_perdoc: Optional[Dict[str, List[Tuple[str, str, Optional[bool], int]]]] = None,
    pipeline_colors: Optional[Dict[str, str]] = None,
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception as e:
        print(f"openpyxl not available ({e}); skipping XLSX export.", file=sys.stderr)
        return

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "CUI_Counts_All"
    headers1 = ["CUI", "Negated", "TotalCount"]
    ws1.append(headers1)
    for cui, neg, cnt in agg_rows:
        ws1.append([cui, neg if neg is not None else "", cnt])

    if per_doc_rows:
        ws2 = wb.create_sheet("Per_Document_All")
        headers2 = ["Document", "CUI", "Negated", "Count"]
        ws2.append(headers2)
        for doc, cui, neg, cnt in per_doc_rows:
            ws2.append([doc, cui, neg if neg is not None else "", cnt])

    if per_pipeline_aggs:
        for pname, agg in sorted(per_pipeline_aggs.items()):
            ws = wb.create_sheet(f"CUI_Counts_{pname}")
            ws.append(headers1)
            for cui, neg, cnt in agg:
                ws.append([cui, neg if neg is not None else "", cnt])
            if per_pipeline_perdoc and pname in per_pipeline_perdoc:
                ws_pd = wb.create_sheet(f"Per_Document_{pname}")
                headers2 = ["Document", "CUI", "Negated", "Count"]
                ws_pd.append(headers2)
                for doc, cui, neg, cnt in per_pipeline_perdoc[pname]:
                    ws_pd.append([doc, cui, neg if neg is not None else "", cnt])

    default_hex = _parse_color_hex(header_color) or "1F4E78"
    header_font = Font(color="FFFFFF", bold=True)
    palette = [
        "00A3E0", "E07A00", "8E44AD", "27AE60", "C0392B", "F1C40F", "2C3E50",
        "16A085", "D35400", "7F8C8D",
    ]
    color_for: Dict[str, str] = {}
    if pipeline_colors:
        for k, v in pipeline_colors.items():
            if _parse_color_hex(v):
                color_for[k] = _parse_color_hex(v)  # type: ignore

    def style_header(ws, hex_color: str):
        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        for cell in ws[1]:
            cell.fill = fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for ws in wb.worksheets:
        if ws.title.startswith("CUI_Counts_All") or ws.title.startswith("Per_Document_All"):
            style_header(ws, default_hex)

    for ws in wb.worksheets:
        m = re.match(r"^(CUI_Counts_|Per_Document_)(.+)$", ws.title)
        if not m:
            continue
        prefix, pname = m.groups()
        if prefix in ("CUI_Counts_", "Per_Document_") and pname not in ("All",):
            hex_color = color_for.get(pname)
            if not hex_color:
                idx = (sum(ord(c) for c in pname) % len(palette))
                hex_color = palette[idx]
            style_header(ws, hex_color)

    # Approximate auto-fit
    def autofit(ws):
        widths: Dict[int, int] = defaultdict(int)
        for row in ws.iter_rows(values_only=True):
            for i, v in enumerate(row, start=1):
                l = len(str(v)) if v is not None else 0
                widths[i] = max(widths[i], l)
        for i, w in widths.items():
            ws.column_dimensions[chr(64 + i)].width = min(80, max(10, int(w * 1.1) + 2))

    for ws in wb.worksheets:
        autofit(ws)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(xlsx_path))


def main():
    p = argparse.ArgumentParser(description="Consolidate cTAKES .cuicount.bsv files into CSV/XLSX")
    p.add_argument("--input-root", required=True, help="Root directory to scan recursively for .cuicount.bsv")
    p.add_argument("--out-base", required=True, help="Output base path (without extension). Creates .xlsx/.csv")
    p.add_argument("--pipeline-name", default=None, help="Pipeline name to label in Excel header")
    p.add_argument("--header-color", default=None, help="Header color hex like #00A3E0 (optional)")
    p.add_argument("--derive-pipeline-from-path", action="store_true", help="Derive pipeline name from file paths; ignored if --pipeline-name is set")
    p.add_argument("--pipeline-colors", default=None, help="Mapping like 'S_core=#00A3E0;S_core_temp_coref=#2C3E50' for per-pipeline colors")
    p.add_argument("--include-per-doc", action="store_true", help="Include per-document sheet and CSV")
    args = p.parse_args()

    root = Path(args.input_root).resolve()
    files = find_cuicount_files(root)
    if not files:
        print(f"No .cuicount.bsv files found under {root}", file=sys.stderr)
        sys.exit(1)

    all_rows: List[CuiRow] = []
    for fp in files:
        try:
            rows = list(read_cuicount_file(fp, root, args.pipeline_name if not args.derive_pipeline_from_path else None))
            all_rows.extend(rows)
        except Exception as e:
            print(f"Failed to read {fp}: {e}", file=sys.stderr)

    if not all_rows:
        print("No rows parsed from cuicount files.", file=sys.stderr)
        sys.exit(2)

    agg_rows = aggregate(all_rows)
    by_pipeline = aggregate_by_pipeline(all_rows)

    out_base = Path(args.out_base)
    # CSV outputs
    write_csv(out_base.with_suffix(".aggregated.csv"), ["CUI", "Negated", "TotalCount"], agg_rows)

    per_doc_rows: Optional[List[Tuple[str, str, Optional[bool], int]]] = None
    per_pipeline_perdoc: Optional[Dict[str, List[Tuple[str, str, Optional[bool], int]]]] = None
    if args.include_per_doc:
        per_doc_rows = sorted(
            [(r.doc, r.cui, r.negated, r.count) for r in all_rows],
            key=lambda x: (x[0], x[1], (1 if x[2] else 0) if x[2] is not None else 2, -x[3]),
        )
        write_csv(out_base.with_suffix(".per_document.csv"), ["Document", "CUI", "Negated", "Count"], per_doc_rows)
        tmp: Dict[str, List[Tuple[str, str, Optional[bool], int]]] = defaultdict(list)
        for r in all_rows:
            tmp[r.pipeline].append((r.doc, r.cui, r.negated, r.count))
        per_pipeline_perdoc = {k: sorted(v, key=lambda x: (x[0], x[1])) for k, v in tmp.items()}

    # Default locked color palette per pipeline (legacy-aligned), can be overridden via --pipeline-colors
    default_mapping: Dict[str, str] = {
        "S_core": "#00A3E0",
        "S_core_rel": "#4F81BD",
        "S_core_temp": "#9BBB59",
        "S_core_temp_coref": "#2C3E50",
        "S_core_temp_coref_smoke": "#8AB4F8",
        "D_core_rel": "#C0504D",
        "D_core_temp": "#8064A2",
        "D_core_temp_coref": "#4BACC6",
        "D_core_temp_coref_smoke": "#2E86C1",
        "WSD_Compare": "#8E44AD",
        "TsSectionedTemporalCoref": "#E07A00",
    }

    mapping: Optional[Dict[str, str]] = dict(default_mapping)
    if args.pipeline_colors:
        # Override defaults with user-provided mapping
        for part in re.split(r"[;,]", args.pipeline_colors):
            part = part.strip()
            if not part:
                continue
            if "=" in part:
                k, v = part.split("=", 1)
                mapping[k.strip()] = v.strip()

    write_xlsx(
        out_base.with_suffix(".xlsx"),
        agg_rows,
        per_doc_rows=per_doc_rows,
        pipeline_name=args.pipeline_name,
        header_color=args.header_color,
        per_pipeline_aggs=by_pipeline,
        per_pipeline_perdoc=per_pipeline_perdoc,
        pipeline_colors=mapping,
    )

    print(
        "Wrote: \n  - "
        + str(out_base.with_suffix(".aggregated.csv"))
        + ("\n  - " + str(out_base.with_suffix(".per_document.csv")) if args.include_per_doc else "")
        + "\n  - "
        + str(out_base.with_suffix(".xlsx"))
    )


if __name__ == "__main__":
    main()
